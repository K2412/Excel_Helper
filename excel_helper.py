import streamlit as st
from openpyxl.styles import Font   
from openpyxl import Workbook, load_workbook
from  tempfile import NamedTemporaryFile


# @st.cache
def file_check(excel1, excel2):
    """takes in two excels and compares the values in the first column. If there are matches it will bold the corresponding value in the first excel"""
    for i in range(2,ws.max_row):
        for y in range(2,bw.max_row):
            if (ws.cell(row=i, column=1).value) == (bw.cell(row=y, column=1).value):
                ws.cell(row=i, column=1).font = Font(bold=True)
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    return stream



st.title('Finding Duplicates')
st.markdown('This app will take two excel files and evaluate duplicates in the first column only. It will then produce a new excel with the duplicates bolded')

data1 = st.file_uploader("Upload the First excel")
data2 = st.file_uploader("Upload the Second excel")

if data1 is not None:
    wb = load_workbook(data1)
    ws = wb.active
if data2 is not None:
    bk = load_workbook(data2)
    bw = bk.active


if st.button('Run The Check'):
    excel = file_check(ws,bw)
    st.download_button(label="Download your data here" , data=excel, file_name="new.xlsx")
    st.balloons()
else:
    st.info("click 'Run The Check'")
